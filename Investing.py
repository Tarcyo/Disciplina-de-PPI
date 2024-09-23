import sys
import os
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import torch
from datetime import datetime
import pytz
from urllib.parse import quote
import google.generativeai as genai
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# Redirecionar stderr para nul para suprimir os erros
sys.stderr = open(os.devnull, 'w')

from selenium import webdriver
from selenium.webdriver.edge.options import Options

edge_options = Options()
edge_options.add_argument('--headless')
edge_options.add_argument('--disable-gpu')
edge_options.add_argument('--no-sandbox')
driver = webdriver.Edge(options=edge_options)

# Carregar o modelo e o tokenizer
model_name = "lucas-leme/FinBERT-PT-BR"
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForSequenceClassification.from_pretrained(model_name)

# Faixa de datas com timezone
brazil_tz = pytz.timezone('America/Sao_Paulo')
start_date = brazil_tz.localize(datetime.strptime("01/06/2024", "%d/%m/%Y"))
end_date = brazil_tz.localize(datetime.strptime("01/08/2024", "%d/%m/%Y"))

# Função para formatar a data para a URL de busca
def format_date_for_url(date):
    return quote(date.strftime("%Y-%m-%dT%H:%M:%S%z"))

def get_news_urls(query):
    url = f"https://br.investing.com/search/?q={quote(query)}&tab=news"
    driver.get(url)

    # Esperar pelo carregamento dos links
    try:
        links = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a.img"))
        )
    except:
        print(f"Nenhum resultado encontrado para a query '{query}'")
        return []

    news_urls = []
    for link in links:
        href = link.get_attribute('href')
        if href.startswith("/"):
            href = "https://br.investing.com" + href
        print(f"URL encontrada: {href}")
        news_urls.append(href)
    
    return news_urls

def clean_text_with_links(element):
    # O Selenium não precisa disso, mas você pode ajustar como preferir
    return element.text

def get_news_text(url):
    driver.get(url)

    # Esperar pelo carregamento do conteúdo principal
    try:
        main_content = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.article_container"))
        )
    except:
        print(f"Não foi possível encontrar o conteúdo principal para a URL '{url}'")
        return None

    # Obter o título da notícia
    try:
        data_hora_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.mt-2.flex.flex-col.gap-2.text-xs.md\\:mt-2\\.5.md\\:gap-2\\.5 div.flex.flex-row.items-center span"))
        )
        data_hora = data_hora_element.text
        publication_date = data_hora
    except:
        print(f"Não foi possível encontrar a data de publicação para a URL '{url}'")
        return None
    
    # Encontrar o título da notícia
    try:
        titulo_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.mx-0.mt-1 h1#articleTitle"))
        )
        titulo = titulo_element.text
        news_title = titulo
    except:
        print("Título da notícia não encontrado")

    
    

    # Obter o texto da notícia
    content_paragraphs = main_content.find_elements(By.CSS_SELECTOR, "p")

    news_content = []
    for paragraph in content_paragraphs:
        cleaned_paragraph = clean_text_with_links(paragraph)
        news_content.append(cleaned_paragraph)
    
    full_content = "\n".join(news_content)
    
    return {
        "title": news_title,
        "publication_date": publication_date,
        "content": full_content
    }

def decimal_to_percentage(decimal_number):
    percentage = decimal_number * 100
    return f"{percentage:.1f}%"

def analyze_sentiment(text):
    inputs = tokenizer(text, return_tensors="pt", truncation=True, padding=False, max_length=512)
    outputs = model(**inputs)
    logits = outputs.logits
    softmax = torch.nn.functional.softmax(logits, dim=-1)
    scores = softmax.cpu().detach().numpy()

    positivo = decimal_to_percentage(scores[0][0])
    negativo = decimal_to_percentage(scores[0][1])
    neutro = decimal_to_percentage(scores[0][2])

    return positivo, negativo, neutro

def resumeOTexto(texto, index):
    chaveUsada = ""
    response = ""
    
    try:
        # API Gemini para resumir o texto:
        # Intercalando com outra chave para evitar bloqueio de segurança por múltiplos acessos:
        if index % 2 == 0:
            genai.configure(api_key="AIzaSyAzvZOfWwkDHOjUdxb8HVveUCA89O6vDPM")
            chaveUsada = "chave da Mariana"
            API = genai.GenerativeModel(model_name="gemini-1.5-flash")
            response = API.generate_content("Retorne para mim apenas uma String que seja um resumo do texto abaixo, ela deve ser curta, com no máximo 512 caracteres mas que contenha todas as informações. Obrigada!\n" + texto)
        else:
            genai.configure(api_key="AIzaSyC_qsbOddcFtSVM1kk2XrMQi_fKJ07fV_Y")
            chaveUsada = "chave do Tarcyo"
            API = genai.GenerativeModel(model_name="gemini-1.5-flash")
            response = API.generate_content("Poderia me fazer um favor? faça um resumo da notícia abaixo por gentileza, garanta que o resumo tenha no máximo 512 caracteres, por favor retorne apenas o texto do resumo. Grato!\n" + texto)
        
        # Verifica se a resposta contém um resumo válido
        if not response or not response.text:
            raise ValueError("Resposta inválida ou bloqueada.")
        
        return response.text, chaveUsada
    
    except Exception as e:
        print(f"Erro ao gerar o resumo: {e}")
        # Retorna um resumo vazio ou alguma string padrão e chave usada
        return "Resumo não disponível devido a um erro.", chaveUsada

def criaAPlanilha(titulos_tabelas, dados_tabelas):
    # Obter o caminho da área de trabalho do usuário
    caminho_area_de_trabalho = os.path.join(os.path.expanduser("~"), "Desktop")

    # Nome do arquivo
    nome_arquivo = "Investing.xlsx"

    # Caminho completo do arquivo na área de trabalho
    caminho_completo = os.path.join(caminho_area_de_trabalho, nome_arquivo)

    # Criar um novo workbook
    wb = Workbook()

    # Remover a planilha padrão criada automaticamente
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Adicionar múltiplas tabelas
    for idx, (titulo_tabela, dados) in enumerate(zip(titulos_tabelas, dados_tabelas), start=1):
        ws = wb.create_sheet(title=titulo_tabela)

        # Adicionar o título da tabela
        ws.merge_cells('A1:F1')
        ws['A1'] = titulo_tabela
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A1'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Criar DataFrame com os dados da tabela
        df = pd.DataFrame(dados)

        # Inserir os dados do DataFrame na planilha
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Salvar o arquivo Excel
    wb.save(caminho_completo)

    print(f"Tabelas criadas com sucesso no arquivo '{nome_arquivo}' na área de trabalho!")        



Titulos=[]
datas=[]
URLs=[]
positivos=[]
negativos=[]
neutros=[]


# Lista de palavras para buscar notícias
queries = ["ELON MUSK", "MARK ZUCKERBERG","LUIZA TRAJANO","BETO SICUPIRA"]

# Obter URLs de notícias para cada palavra na lista
for i in range(len(queries)):
    query=queries[i]
    print(f"RESULTADOS PARA: '{query}':")
    print()
    Titulos.append([])
    datas.append([])
    URLs.append([])
    positivos.append([])
    negativos.append([])
    neutros.append([])
    news_urls = get_news_urls(query)
    if not news_urls:
        continue
    for j in range(len(news_urls)):
        url=news_urls[j]
        news_data = get_news_text(url)
        if news_data:
            try:
                resumo, chaveUsada = resumeOTexto(news_data['content'], j)
                positivo, negativo, neutro = analyze_sentiment(resumo)
                
                print(f"Título: {news_data['title']}")
                print(f"Data de Publicação: {news_data['publication_date']}")
                print(f"Resumo do texto: {resumo}")
                print(f"Chave usada: {chaveUsada}")
                print(f"URL: {url}")
                print(f"SENTIMENTOS: Positivo: {positivo} Negativo: {negativo} Neutro: {neutro}")
                print()
                print()

                Titulos[i].append(news_data['title'])
                datas[i].append(news_data['publication_date'])
                URLs[i].append(url)
                positivos[i].append(positivo)
                negativos[i].append(negativo)
                neutros[i].append(neutro)
            except Exception as e:
                print(f"Erro ao processar a notícia '{url}': {e}")
                continue  # Continua para a próxima URL em caso de erro

# Fechar o navegador
driver.quit()



print("AQUI")
dados_tabelas = [
]

for numeroQuery in range(len(queries)):
    dados_tabelas.append({"Título": Titulos[numeroQuery],
        "Data": datas[numeroQuery],
        "URL": URLs[numeroQuery],
        "Positivo": positivos[numeroQuery],
        "Negativo": negativos[numeroQuery],
        "Neutro": neutros[numeroQuery]})
    




criaAPlanilha(queries,dados_tabelas)


