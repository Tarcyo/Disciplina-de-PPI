import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill

# Dados para as colunas, fornecidos por vetores
titulos = ["Notícia 1", "Notícia 2", "Notícia 3"]
datas = ["2024-08-18", "2024-08-17", "2024-08-16"]
urls = ["http://exemplo.com/noticia1", "http://exemplo.com/noticia2", "http://exemplo.com/noticia3"]
positivos = [0.75, 0.60, 0.20]
negativos = [0.10, 0.25, 0.70]
neutros = [0.15, 0.15, 0.10]

# Criação do DataFrame com os dados
df = pd.DataFrame({
    "Título": titulos,
    "Data": datas,
    "URL": urls,
    "Positivo": positivos,
    "Negativo": negativos,
    "Neutro": neutros
})

# Obter o caminho da área de trabalho do usuário
caminho_area_de_trabalho = os.path.join(os.path.expanduser("~"), "Desktop")

# Nome do arquivo
nome_arquivo = "Sentimentos_das_Notícias_Declarações.xlsx"

# Caminho completo do arquivo na área de trabalho
caminho_completo = os.path.join(caminho_area_de_trabalho, nome_arquivo)

# Criar um novo workbook e selecionar a active sheet
wb = Workbook()
ws = wb.active

# Mesclar as células para o título e definir o texto
ws.merge_cells('A1:F1')
ws['A1'] = "Sentimentos das notícias/declarações"
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Definir a cor de fundo do título como verde
ws['A1'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Inserir os dados do DataFrame na planilha
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Salvar o arquivo Excel
wb.save(caminho_completo)

print(f"Tabela '{nome_arquivo}' criada com sucesso na área de trabalho!")
