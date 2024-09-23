import time
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pytz
from urllib.parse import urlparse, parse_qs, quote
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import torch
import google.generativeai as genai






# Carregar o modelo e o tokenizer
model_name = "lucas-leme/FinBERT-PT-BR"
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForSequenceClassification.from_pretrained(model_name)

# Faixa de datas com timezone
brazil_tz = pytz.timezone('America/Sao_Paulo')
start_date = brazil_tz.localize(datetime.strptime("01/08/2024", "%d/%m/%Y"))
end_date = brazil_tz.localize(datetime.strptime("23/09/2024", "%d/%m/%Y"))

# Função para formatar a data para a URL de busca
def format_date_for_url(date):
    return quote(date.strftime("%Y-%m-%dT%H:%M:%S%z"))

def get_news_urls(query):
    # Formatar a URL de busca com a query fornecida e a faixa de datas
    from_date = format_date_for_url(start_date)
    to_date = format_date_for_url(end_date)
    url = f"https://g1.globo.com/busca/?q={query}&ps=on&order=recent&species=not%C3%ADcias&from={from_date}&to={to_date}"
    
    # Enviar uma solicitação GET para a URL
    response = requests.get(url)
    
    # Verificar se a solicitação foi bem-sucedida
    if response.status_code != 200:
        print(f"Erro ao acessar a página: {response.status_code}")
        return []
    
    # Parsear o conteúdo HTML da página
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Encontrar todos os elementos que contêm os URLs das notícias
    results_content = soup.find('div', class_='results__content')
    if not results_content:
        print(f"Nenhum resultado encontrado para a query '{query}'")
        return []
    
    news_items = results_content.find_all('div', class_='widget--info__text-container')
    
    news_urls = []
    for item in news_items:
        a_tag = item.find('a')
        if a_tag:
            link = a_tag['href']
            if link.startswith("//"):
                link = "https:" + link
            # Extrair a URL real do parâmetro `u` se a URL contiver `click?q=`
            parsed_url = urlparse(link)
            if 'click' in parsed_url.path:
                query_params = parse_qs(parsed_url.query)
                if 'u' in query_params:
                    link = query_params['u'][0]
            news_urls.append(link)
    
    return news_urls

def convert_to_brazil_timezone(date_str):
    # Remover fração de segundos se presente
    if '.' in date_str:
        date_str = date_str.split('.')[0] + 'Z'
    utc_time = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%SZ")
    utc_time = utc_time.replace(tzinfo=pytz.utc)
    brazil_time = utc_time.astimezone(brazil_tz)
    return brazil_time

def clean_text_with_links(element):
    # Adicionar espaços antes e depois de links clicáveis
    for a in element.find_all('a', attrs={"cmp-ltrk": "Article links"}):
        a.insert_before(" ")
        a.insert_after(" ")
    return element

def get_news_text(url):
    # Enviar uma solicitação GET para a URL da notícia
    response = requests.get(url)
    
    # Verificar se a solicitação foi bem-sucedida
    if response.status_code != 200:
        print(f"Erro ao acessar a notícia: {response.status_code}")
        return None
    
    # Parsear o conteúdo HTML da página
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Encontrar o título da notícia
    main_content = soup.find('main', class_='mc-body theme')
    if not main_content:
        print(f"Não foi possível encontrar o conteúdo principal para a URL '{url}'")
        return None
    
    # Obter o título da notícia
    title_tag = main_content.find('h1', class_='content-head__title', itemprop='headline')
    if not title_tag:
        print(f"Não foi possível encontrar o título para a URL '{url}'")
        return None
    news_title = title_tag.get_text(strip=True)
    
    # Obter a data e hora da publicação
    time_tag = main_content.find('time', itemprop='datePublished')
    if not time_tag:
        print(f"Não foi possível encontrar a data de publicação para a URL '{url}'")
        return None
    date_str = time_tag['datetime']
    publication_date = convert_to_brazil_timezone(date_str)
    
    # Verificar se a data está dentro da faixa desejada
    if publication_date < start_date or publication_date > end_date:
        return None
    
    # Obter o texto da notícia
    content_paragraphs = main_content.find_all('p', class_='content-text__container')
    content_divs = main_content.find_all('div', class_='wall protected-content')
    
    news_content = []
    for paragraph in content_paragraphs:
        cleaned_paragraph = clean_text_with_links(paragraph)
        news_content.append(cleaned_paragraph.get_text(" ", strip=True))
    
    for div in content_divs:
        cleaned_div = clean_text_with_links(div)
        news_content.append(cleaned_div.get_text(" ", strip=True))
    
    full_content = "\n".join(news_content)
    
    return {
        "title": news_title,
        "publication_date": publication_date.strftime("%d/%m/%Y %H:%M"),
        "content": full_content
    }


def decimal_to_percentage(decimal_number):
    percentage = decimal_number * 100
    return f"{percentage:.1f}%"



def analyze_sentiment(text):
    inputs = tokenizer(text, return_tensors="pt", truncation=True, padding=False, max_length=512)
    outputs = model(**inputs)
    logits=outputs.logits
    softmax = torch.nn.functional.softmax(logits, dim=-1)
    scores = softmax.cpu().detach().numpy()

    positivo=decimal_to_percentage(scores[0][0])
    negativo=decimal_to_percentage(scores[0][1])
    neutro=decimal_to_percentage(scores[0][2])

    return positivo,negativo, neutro

def resumeOTexto(texto, index):
    chaveUsada = ""
    response = ""
    
    try:
        # API Gemini para resumir o texto:
        # Intercalando com outra chave para evitar bloqueio de segurança por múltiplos acessos:
        if index % 2 == 0:
            genai.configure(api_key="AIzaSyAzvZOfWwkDHOjUdxb8HVveUCA89O6vDPM")
            chaveUsada = "chave da Mariana"
            API = genai.GenerativeModel(model_name="gemini-1.5-flash")
            response = API.generate_content("Retorne para mim apenas uma String que seja um resumo do texto abaixo, ela deve ser curta, com no máximo 512 caracteres mas que contenha todas as informações. Obrigada!\n" + texto)
        else:
            genai.configure(api_key="AIzaSyC_qsbOddcFtSVM1kk2XrMQi_fKJ07fV_Y")
            chaveUsada = "chave do Tarcyo"
            API = genai.GenerativeModel(model_name="gemini-1.5-flash")
            response = API.generate_content("Poderia me fazer um favor? faça um resumo da notícia abaixo por gentileza, garanta que o resumo tenha no máximo 512 caracteres, por favor retorne apenas o texto do resumo. Grato!\n" + texto)
        
        # Verifica se a resposta contém um resumo válido
        if not response or not response.text:
            raise ValueError("Resposta inválida ou bloqueada.")
        
        return response.text, chaveUsada
    
    except Exception as e:
        print(f"Erro ao gerar o resumo: {e}")
        # Retorna um resumo vazio ou alguma string padrão e chave usada
        return "Resumo não disponível devido a um erro.", chaveUsada


def criaAPlanilha(titulos_tabelas,dados_tabelas):

        # Obter o caminho da área de trabalho do usuário
        caminho_area_de_trabalho = os.path.join(os.path.expanduser("~"), "Desktop")

        # Nome do arquivo
        nome_arquivo = "Sentimentos_das_Notícias_Declarações.xlsx"

        # Caminho completo do arquivo na área de trabalho
        caminho_completo = os.path.join(caminho_area_de_trabalho, nome_arquivo)

        # Criar um novo workbook
        wb = Workbook()

        # Remover a planilha padrão criada automaticamente
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Adicionar múltiplas tabelas
        for idx, (titulo_tabela, dados) in enumerate(zip(titulos_tabelas, dados_tabelas), start=1):
            ws = wb.create_sheet(title=titulos_tabelas[idx-1])

            # Adicionar o título da tabela
            ws.merge_cells('A1:F1')
            ws['A1'] = titulo_tabela
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A1'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            # Criar DataFrame com os dados da tabela
            df = pd.DataFrame(dados)

            # Inserir os dados do DataFrame na planilha
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Salvar o arquivo Excel
        wb.save(caminho_completo)

        print(f"Tabelas criadas com sucesso no arquivo '{nome_arquivo}' na área de trabalho!")
        
        


# Lista de palavras para buscar notícias
queries = ["ELON MUSK", "MARK ZUCKERBERG",]




Titulos=[]
datas=[]
URLs=[]
positivos=[]
negativos=[]
neutros=[]





# Obter URLs de notícias para cada palavra na lista
for numeroQuery in range(len(queries)):
    query = queries[numeroQuery]
    print(f"                                                   RESULTADOS PARA: '{query}':")
    print()
    news_urls = get_news_urls(query)

    Titulos.append([])
    datas.append([])
    URLs.append([])
    positivos.append([])
    negativos.append([])
    neutros.append([])

    if not news_urls:
        continue

    for numeroUrl in range(len(news_urls)):
        # Pausando para evitar o bloqueio da API por sobrecarga
        time.sleep(4)

        url = news_urls[numeroUrl]
        news_data = get_news_text(url)
        if news_data:
            try:
                resumo, chaveUsada = resumeOTexto(news_data['content'], numeroUrl)
                positivo, negativo, neutro = analyze_sentiment(resumo)
                
                print(f"Título: {news_data['title']}")
                print(f"Data de Publicação: {news_data['publication_date']}")
                print(f"Resumo do texto: {resumo}")
                print(f"Chave usada: {chaveUsada}")
                print(f"URL: {url}")
                print(f"SENTIMENTOS: Positivo: {positivo} Negativo: {negativo} Neutro: {neutro}")
                print()
                print()

                Titulos[numeroQuery].append(news_data['title'])
                datas[numeroQuery].append(news_data['publication_date'])
                URLs[numeroQuery].append(url)
                positivos[numeroQuery].append(positivo)
                negativos[numeroQuery].append(negativo)
                neutros[numeroQuery].append(neutro)
            except Exception as e:
                print(f"Erro ao processar a notícia '{url}': {e}")
                continue  # Continua para a próxima URL em caso de erro

dados_tabelas = [
]

for numeroQuery in range(len(queries)):
    dados_tabelas.append({"Título": Titulos[numeroQuery],
        "Data": datas[numeroQuery],
        "URL": URLs[numeroQuery],
        "Positivo": positivos[numeroQuery],
        "Negativo": negativos[numeroQuery],
        "Neutro": neutros[numeroQuery]})
    




criaAPlanilha(queries,dados_tabelas)



            
    

   
